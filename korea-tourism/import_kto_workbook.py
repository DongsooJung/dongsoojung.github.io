#!/usr/bin/env python3
"""한국관광 데이터랩의 최신 공식 엑셀을 대시보드 JSON으로 변환한다."""

from __future__ import annotations

import argparse
import json
import re
import tempfile
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
MAIN_URL = "https://datalab.visitkorea.or.kr/datalab/portal/main/getMainForm.do"
COUNTRIES = {"china": "중국", "taiwan": "대만", "vietnam": "베트남"}


def latest_workbook_url() -> str:
    req = urllib.request.Request(MAIN_URL, headers={"User-Agent": "stargateedu-dashboard/1.0"})
    html = urllib.request.urlopen(req, timeout=45).read().decode("utf-8", "replace")
    marker = re.search(r"20\d{2}년\s*\d{1,2}월\s*한국관광통계\s*공표", html)
    if not marker:
        raise RuntimeError("한국관광통계 공표 게시물을 찾지 못했습니다.")
    window = html[marker.start() : marker.start() + 6000]
    match = re.search(
        r"(?:href=)?[\"']?\s*(/common/board/Download\.do\?[^\"'<>\s]+\.xlsx)", window
    )
    if not match:
        raise RuntimeError("공식 XLSX 다운로드 링크를 찾지 못했습니다.")
    return urllib.parse.urljoin(MAIN_URL, match.group(1).replace("&amp;", "&"))


def obtain(source: str | None) -> tuple[Path, str]:
    if source and Path(source).exists():
        return Path(source), "local-official-workbook"
    url = source or latest_workbook_url()
    target = Path(tempfile.gettempdir()) / "kto-tourism-latest.xlsx"
    req = urllib.request.Request(url, headers={"User-Agent": "stargateedu-dashboard/1.0"})
    with urllib.request.urlopen(req, timeout=90) as response:
        target.write_bytes(response.read())
    return target, url


def parse(path: Path, source_url: str, start_year: int) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "입국" not in workbook.sheetnames:
        raise RuntimeError("공식 통계 파일의 '입국' 시트가 없습니다.")
    sheet = workbook["입국"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
    month_cols: dict[int, str] = {}
    for column, value in enumerate(headers, start=1):
        match = re.fullmatch(r"(20\d{2})\.(\d{1,2})월", str(value or "").strip())
        if match and int(match.group(1)) >= start_year:
            month_cols[column] = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    if not month_cols:
        raise RuntimeError("월별 통계 열을 찾지 못했습니다.")

    country_rows: dict[str, dict[str, int | None]] = {}
    label_to_key = {label: key for key, label in COUNTRIES.items()}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        label = str(row[1] or "").strip()
        if label not in label_to_key:
            continue
        key = label_to_key[label]
        country_rows[key] = {
            ym: int(row[column - 1]) if isinstance(row[column - 1], (int, float)) else None
            for column, ym in month_cols.items()
        }
    missing = set(COUNTRIES) - set(country_rows)
    if missing:
        raise RuntimeError(f"국가 행 누락: {', '.join(sorted(missing))}")

    series = []
    for ym in sorted(month_cols.values()):
        item = {"ym": ym, **{key: country_rows[key][ym] for key in COUNTRIES}}
        if any(item[key] is not None for key in COUNTRIES):
            series.append(item)
    if not series or series[-1]["ym"] < "2024-01":
        raise RuntimeError("공식 파일의 최신 월 검증에 실패했습니다.")

    return {
        "updatedAt": datetime.now(timezone.utc).date().isoformat(),
        "sourceMode": "official-xlsx",
        "source": "한국관광공사 한국관광 데이터랩 한국관광통계 공식 XLSX",
        "sourceUrl": source_url,
        "series": series,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", nargs="?", help="공식 XLSX 로컬 경로 또는 다운로드 URL")
    parser.add_argument("--start-year", type=int, default=2010)
    args = parser.parse_args()
    workbook, source_url = obtain(args.source)
    data = parse(workbook, source_url, args.start_year)
    encoded = json.dumps(data, ensure_ascii=False, indent=2) + "\n"
    (ROOT / "data.json").write_text(encoded, encoding="utf-8")
    (ROOT / "fallback-data.js").write_text(
        "window.FALLBACK_DATA = " + json.dumps(data, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(
        f"공식 통계 {len(data['series'])}개월 변환 완료: "
        f"{data['series'][0]['ym']} ~ {data['series'][-1]['ym']}"
    )


if __name__ == "__main__":
    main()
