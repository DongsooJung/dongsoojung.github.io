#!/usr/bin/env python3
"""DRAM·브렌트유·금·구리·천연가스를 연간 평균으로 집계한다.

DRAM은 Stanford DAM 공개 CSV의 전체 DRAM 최저 소매가격(USD/GB)을 사용한다.
2024년 7월까지 McCallum 계열, 2024년 8월부터 Keepa 계열을 이어 붙인다.
유가는 FRED가 배포하는 EIA 브렌트유 일일 현물가격(DCOILBRENTEU)을 사용한다.
금·구리·미국 천연가스는 세계은행 Pink Sheet 월별 가격을 사용한다.

두 소스 모두 인증키 없이 내려받을 수 있다.
"""

import csv
import io
import json
import sys
import urllib.request
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path


# 환율·금리 커버리지(1999~)와 맞추고, 원자재·유가는 그 이전부터도 가능하다.
START_YEAR = 1999
HTTP_TIMEOUT = 90
DRAM_URL = "https://dam.stanford.edu/assets/memory-prices/memory-prices.csv"
BRENT_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=DCOILBRENTEU"
WORLD_BANK_URL = "https://thedocs.worldbank.org/en/doc/74e8be41ceb20fa0da750cda2f6b9e4e-0050012026/related/CMO-Historical-Data-Monthly.xlsx"
DATA_PATH = Path(__file__).parent / "market_data.json"
INDEX_BASE_YEAR = 1999  # 원자재 가격지수 기준 연도


def download_bytes(url):
    req = urllib.request.Request(url, headers={"User-Agent": "stargateedu-dashboard/1.0"})
    with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT) as response:
        return response.read()


def download_text(url):
    return download_bytes(url).decode("utf-8-sig")


def mean(values):
    return sum(values) / len(values) if values else None


def fetch_dram():
    rows = csv.DictReader(io.StringIO(download_text(DRAM_URL)))
    yearly = defaultdict(list)
    last_date = None
    for row in rows:
        if row.get("category") != "DRAM" or row.get("metric") != "usd_per_gb":
            continue
        observed = row.get("date", "")
        try:
            year = int(observed[:4])
            value = float(row["value"])
        except (ValueError, TypeError, KeyError):
            continue
        if year < START_YEAR:
            continue
        series = row.get("series", "")
        use_historical = series == "McCallum DRAM (historical)" and observed < "2024-08-01"
        use_live = series == "DRAM cheapest (Keepa)" and observed >= "2024-08-01"
        if not (use_historical or use_live):
            continue
        yearly[year].append(value)
        if last_date is None or observed > last_date:
            last_date = observed
    if not yearly:
        raise RuntimeError("Stanford DAM DRAM 데이터가 비어 있습니다.")
    return yearly, last_date


def fetch_brent():
    rows = csv.DictReader(io.StringIO(download_text(BRENT_URL)))
    yearly = defaultdict(list)
    last_date = None
    for row in rows:
        observed = row.get("observation_date", "")
        try:
            year = int(observed[:4])
            value = float(row["DCOILBRENTEU"])
        except (ValueError, TypeError, KeyError):
            continue
        if year < START_YEAR:
            continue
        yearly[year].append(value)
        if last_date is None or observed > last_date:
            last_date = observed
    if not yearly:
        raise RuntimeError("FRED 브렌트유 데이터가 비어 있습니다.")
    return yearly, last_date


def fetch_world_bank():
    """세계은행 Pink Sheet에서 금·구리·미국 천연가스·브렌트유 월가격을 읽는다."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl 설치가 필요합니다.") from exc

    workbook = openpyxl.load_workbook(
        io.BytesIO(download_bytes(WORLD_BANK_URL)), read_only=True, data_only=True
    )
    sheet = workbook["Monthly Prices"]
    headers = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
    columns = {
        "gasUsdPerMmbtu": headers.index("Natural gas, US"),
        "copperUsdPerMt": headers.index("Copper"),
        "goldUsdPerOz": headers.index("Gold"),
        "brentUsdPerBbl": headers.index("Crude oil, Brent"),
    }
    yearly = {key: defaultdict(list) for key in columns}
    last_period = None
    for row in sheet.iter_rows(min_row=7, values_only=True):
        period = str(row[0] or "")
        if len(period) < 7 or "M" not in period:
            continue
        try:
            year = int(period[:4])
            month = int(period[-2:])
        except ValueError:
            continue
        if year < START_YEAR:
            continue
        observed = f"{year:04d}-{month:02d}-01"
        for key, column in columns.items():
            value = row[column]
            if isinstance(value, (int, float)):
                yearly[key][year].append(float(value))
        if last_period is None or observed > last_period:
            last_period = observed
    workbook_updated = str(sheet["A4"].value or "").replace("Updated on ", "")
    commodity_keys = ("gasUsdPerMmbtu", "copperUsdPerMt", "goldUsdPerOz")
    if not any(yearly[key] for key in commodity_keys):
        raise RuntimeError("세계은행 원자재 데이터가 비어 있습니다.")
    return yearly, last_period, workbook_updated


def load_existing():
    if not DATA_PATH.exists():
        return {}
    try:
        return json.loads(DATA_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def main():
    existing = load_existing()
    dram, dram_last = {}, None
    brent, brent_last = {}, None
    commodities = {
        "goldUsdPerOz": defaultdict(list),
        "copperUsdPerMt": defaultdict(list),
        "gasUsdPerMmbtu": defaultdict(list),
    }
    commodity_last, workbook_updated = None, None

    try:
        dram, dram_last = fetch_dram()
    except Exception as exc:
        print(f"DRAM 갱신 실패 — 기존 값 유지: {exc}", file=sys.stderr)
    try:
        brent, brent_last = fetch_brent()
    except Exception as exc:
        print(f"브렌트유 갱신 실패 — 기존 값 유지: {exc}", file=sys.stderr)
    try:
        commodities, commodity_last, workbook_updated = fetch_world_bank()
    except Exception as exc:
        print(f"원자재 갱신 실패 — 기존 값 유지: {exc}", file=sys.stderr)

    # 일부 소스만 실패해도 나머지와 기존 JSON을 합친다.
    existing_by_year = {
        int(row["year"]): row for row in existing.get("series", []) if "year" in row
    }
    years = sorted(
        set(dram)
        | set(brent)
        | set(commodities.get("goldUsdPerOz", {}))
        | set(commodities.get("copperUsdPerMt", {}))
        | set(commodities.get("gasUsdPerMmbtu", {}))
        | set(existing_by_year)
    )
    years = [y for y in years if y >= START_YEAR]
    if not years:
        print("시장 데이터가 비어 있어 기존 파일을 유지합니다.", file=sys.stderr)
        return 0

    current_year = date.today().year
    series = []
    for year in years:
        prev = existing_by_year.get(year, {})
        row = {"year": year, "partial": year == current_year}
        if dram.get(year):
            row["dramUsdPerGb"] = round(mean(dram[year]), 3)
            row["dramObservations"] = len(dram[year])
        elif prev.get("dramUsdPerGb") is not None:
            row["dramUsdPerGb"] = prev["dramUsdPerGb"]
            if prev.get("dramObservations") is not None:
                row["dramObservations"] = prev["dramObservations"]
        # 브렌트유: FRED 우선, 없으면 세계은행 Pink Sheet, 그래도 없으면 기존 JSON
        if brent.get(year):
            row["brentUsdPerBbl"] = round(mean(brent[year]), 2)
            row["brentObservations"] = len(brent[year])
        elif commodities.get("brentUsdPerBbl", {}).get(year):
            vals = commodities["brentUsdPerBbl"][year]
            row["brentUsdPerBbl"] = round(mean(vals), 2)
            row["brentObservations"] = len(vals)
        elif prev.get("brentUsdPerBbl") is not None:
            row["brentUsdPerBbl"] = prev["brentUsdPerBbl"]
            if prev.get("brentObservations") is not None:
                row["brentObservations"] = prev["brentObservations"]
        observation_keys = {
            "goldUsdPerOz": "goldObservations",
            "copperUsdPerMt": "copperObservations",
            "gasUsdPerMmbtu": "gasObservations",
        }
        for key, observation_key in observation_keys.items():
            values = commodities.get(key, {}).get(year, [])
            if values:
                row[key] = round(mean(values), 2)
                row[observation_key] = len(values)
            elif prev.get(key) is not None:
                row[key] = prev[key]
                if prev.get(observation_key) is not None:
                    row[observation_key] = prev[observation_key]
        series.append(row)

    old_sources = existing.get("sources", {})
    payload = {
        "updatedAt": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "sourceMode": "public-csv",
        "coverage": {
            "start": series[0]["year"],
            "end": series[-1]["year"],
            "years": len(series),
            "indexBaseYear": INDEX_BASE_YEAR,
        },
        "sources": {
            "dram": {
                "name": "Stanford DAM Memory Prices",
                "metric": "DRAM 최저 소비자 소매가격",
                "unit": "USD/GB",
                "url": "https://dam.stanford.edu/memory-prices.html",
                "observedAt": dram_last or old_sources.get("dram", {}).get("observedAt"),
            },
            "brent": {
                "name": "FRED DCOILBRENTEU / World Bank Pink Sheet Brent",
                "metric": "Brent Europe 현물·월별 가격",
                "unit": "USD/barrel",
                "url": "https://www.worldbank.org/en/research/commodity-markets",
                "observedAt": brent_last
                or commodity_last
                or old_sources.get("brent", {}).get("observedAt"),
            },
            "commodities": {
                "name": "World Bank Commodity Price Data (Pink Sheet)",
                "metrics": "Gold · Copper · Natural gas, US · Crude oil, Brent",
                "url": "https://www.worldbank.org/en/research/commodity-markets",
                "observedAt": commodity_last or old_sources.get("commodities", {}).get("observedAt"),
                "workbookUpdated": workbook_updated
                or old_sources.get("commodities", {}).get("workbookUpdated"),
                "indexBaseYear": INDEX_BASE_YEAR,
            },
        },
        "series": series,
    }
    DATA_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (DATA_PATH.parent / "market-fallback.js").write_text(
        "window.MARKET_FALLBACK = " + json.dumps(payload, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(
        f"완료: DRAM·브렌트유·금·구리·천연가스 {len(series)}개 연도 갱신 "
        f"({series[0]['year']}~{series[-1]['year']})."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
