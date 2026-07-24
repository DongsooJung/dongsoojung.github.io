#!/usr/bin/env python3
"""Archive the public Inflearn homepage course list as JSON and Excel."""

from __future__ import annotations

import argparse
import gzip
import json
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

SOURCE_URL = "https://www.inflearn.com/ko/"
SEOUL = ZoneInfo("Asia/Seoul")
LEVELS = {
    "BEGINNER": "입문",
    "BASIC": "초급",
    "NORMAL": "중급",
    "ADVANCED": "고급",
}


def fetch_homepage() -> str:
    request = Request(
        SOURCE_URL,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (compatible; InflearnDailyArchive/1.0; "
                "+https://stargateedu.co.kr/inflearn/)"
            ),
            "Accept": "text/html,application/xhtml+xml",
            "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.5",
            "Accept-Encoding": "gzip",
        },
    )
    with urlopen(request, timeout=40) as response:
        payload = response.read()
        if response.headers.get("Content-Encoding") == "gzip":
            payload = gzip.decompress(payload)
        charset = response.headers.get_content_charset() or "utf-8"
        return payload.decode(charset, errors="replace")


def extract_course_items(html: str) -> list[dict]:
    match = re.search(
        r'<script[^>]+id=["\']__NEXT_DATA__["\'][^>]*>(.*?)</script>',
        html,
        re.DOTALL,
    )
    if not match:
        raise RuntimeError("인프런 페이지에서 공개 강의 데이터를 찾지 못했습니다.")

    next_data = json.loads(match.group(1))
    queries = (
        next_data.get("props", {})
        .get("pageProps", {})
        .get("dehydratedState", {})
        .get("queries", [])
    )
    for query in queries:
        query_key = query.get("queryKey") or []
        if query_key and query_key[0] == "/client/api/v2/courses/search":
            items = (
                query.get("state", {})
                .get("data", {})
                .get("data", {})
                .get("items", [])
            )
            if items:
                return items
    raise RuntimeError("인프런 메인 강의 목록이 비어 있습니다.")


def normalize_course(item: dict, rank: int, collected_at: str) -> dict:
    course = item.get("course") or {}
    instructor = item.get("instructor") or {}
    price = item.get("listPrice") or {}
    metadata = course.get("metadata") or {}
    parent_categories = metadata.get("parentCategories") or []
    child_categories = metadata.get("childCategories") or []
    skills = metadata.get("skills") or []
    course_id = course.get("id") or item.get("id")
    slug = course.get("slug") or course_id

    regular_price = price.get("regularPrice")
    pay_price = price.get("payPrice")
    is_free = bool(price.get("isFree"))
    if is_free:
        price_label = "무료"
    elif isinstance(pay_price, (int, float)):
        price_label = f"₩{pay_price:,.0f}"
    else:
        price_label = "가격 비공개"

    runtime_seconds = course.get("runtimeSecond")
    return {
        "rank": rank,
        "courseId": course_id,
        "title": course.get("title") or "제목 없음",
        "instructor": instructor.get("name") or "강사 정보 없음",
        "price": price_label,
        "payPrice": pay_price,
        "regularPrice": regular_price,
        "discountRate": price.get("discountRate") or 0,
        "rating": course.get("star"),
        "reviewCount": course.get("reviewCount") or 0,
        "studentCount": course.get("studentCount") or 0,
        "level": LEVELS.get(metadata.get("level"), metadata.get("level") or "미분류"),
        "categories": parent_categories,
        "subCategories": child_categories,
        "skills": skills,
        "runtimeMinutes": round(runtime_seconds / 60) if runtime_seconds else None,
        "publishedAt": course.get("publishedAtUTC") or course.get("publishedAt"),
        "updatedAt": course.get("updatedAtUTC") or course.get("updatedAt"),
        "collectedAt": collected_at,
        "url": f"https://www.inflearn.com/course/{slug}",
        "thumbnailUrl": course.get("thumbnailUrl"),
    }


def save_excel(path: Path, records: list[dict], collected_at: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("엑셀 생성을 위해 openpyxl이 필요합니다.") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "강의목록"

    headers = [
        "순위",
        "강의명",
        "강사",
        "현재가",
        "정가",
        "할인율",
        "평점",
        "후기 수",
        "수강생 수",
        "난이도",
        "상위 카테고리",
        "세부 카테고리",
        "기술 태그",
        "강의 시간(분)",
        "게시 일시",
        "수정 일시",
        "수집 일시",
        "강의 URL",
    ]
    sheet.append(headers)

    for record in records:
        sheet.append(
            [
                record["rank"],
                record["title"],
                record["instructor"],
                record["payPrice"],
                record["regularPrice"],
                record["discountRate"] / 100,
                record["rating"],
                record["reviewCount"],
                record["studentCount"],
                record["level"],
                ", ".join(record["categories"]),
                ", ".join(record["subCategories"]),
                ", ".join(record["skills"]),
                record["runtimeMinutes"],
                record["publishedAt"],
                record["updatedAt"],
                record["collectedAt"],
                record["url"],
            ]
        )

    green = "14C786"
    dark = "17332D"
    pale = "E8F8F1"
    border = Side(style="thin", color="DDE7E3")
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=border)

    for row in range(2, sheet.max_row + 1):
        fill = PatternFill("solid", fgColor=pale if row % 2 == 0 else "FFFFFF")
        for cell in sheet[row]:
            cell.fill = fill
            cell.border = Border(bottom=border)
            cell.alignment = Alignment(vertical="top")
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.cell(row, 4).number_format = '₩#,##0'
        sheet.cell(row, 5).number_format = '₩#,##0'
        sheet.cell(row, 6).number_format = "0%"
        link_cell = sheet.cell(row, 18)
        link_cell.hyperlink = link_cell.value
        link_cell.style = "Hyperlink"
        link_cell.font = Font(color=green, underline="single")

    widths = [7, 48, 20, 13, 13, 10, 9, 11, 12, 12, 24, 28, 34, 14, 24, 24, 24, 54]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26

    info = workbook.create_sheet("수집정보")
    info.append(["항목", "내용"])
    info.append(["데이터 출처", SOURCE_URL])
    info.append(["수집 일시", collected_at])
    info.append(["수집 강의 수", len(records)])
    info.append(["수집 범위", "인프런 비로그인 메인 페이지의 공개 강의 목록"])
    info.append(["안내", "인프런과 무관한 비공식 아카이브이며, 원문은 강의 URL에서 확인하세요."])
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 82
    for cell in info[1]:
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(color="FFFFFF", bold=True)
    info.freeze_panes = "A2"

    workbook.save(path)


def write_outputs(root: Path, records: list[dict], collected_at: str) -> None:
    archive_date = collected_at[:10]
    data_dir = root / "data"
    history_dir = data_dir / "history"
    output_dir = root / "output"
    history_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    snapshot = {
        "source": SOURCE_URL,
        "collectedAt": collected_at,
        "archiveDate": archive_date,
        "count": len(records),
        "courses": records,
    }
    encoded = json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n"
    (data_dir / "latest.json").write_text(encoded, encoding="utf-8")
    (history_dir / f"{archive_date}.json").write_text(encoded, encoding="utf-8")

    dated_excel = output_dir / f"inflearn_main_{archive_date}.xlsx"
    save_excel(dated_excel, records, collected_at)
    shutil.copy2(dated_excel, output_dir / "inflearn_main_latest.xlsx")

    archive_path = data_dir / "archives.json"
    if archive_path.exists():
        archives = json.loads(archive_path.read_text(encoding="utf-8"))
    else:
        archives = {"archives": []}
    entries = [
        item for item in archives.get("archives", []) if item.get("date") != archive_date
    ]
    entries.append(
        {
            "date": archive_date,
            "collectedAt": collected_at,
            "count": len(records),
            "excel": f"./output/inflearn_main_{archive_date}.xlsx",
            "json": f"./data/history/{archive_date}.json",
        }
    )
    archives["archives"] = sorted(entries, key=lambda item: item["date"], reverse=True)
    archive_path.write_text(
        json.dumps(archives, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-html",
        type=Path,
        help="네트워크 대신 저장된 인프런 HTML을 사용합니다.",
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="데이터와 엑셀을 저장할 inflearn 디렉터리",
    )
    args = parser.parse_args()

    html = (
        args.input_html.read_text(encoding="utf-8")
        if args.input_html
        else fetch_homepage()
    )
    now = datetime.now(SEOUL).replace(microsecond=0)
    collected_at = now.isoformat()
    items = extract_course_items(html)
    records = [
        normalize_course(item, rank, collected_at)
        for rank, item in enumerate(items, start=1)
    ]
    if len(records) < 5:
        raise RuntimeError(f"수집된 강의가 너무 적습니다: {len(records)}개")
    write_outputs(args.root.resolve(), records, collected_at)
    print(f"{collected_at} 기준 강의 {len(records)}개를 저장했습니다.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"수집 실패: {error}", file=sys.stderr)
        raise
