#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
법원경매 2026년 월별 부동산(주택/상업용) 데이터 수집기
======================================================

대한민국법원 법원경매정보(www.courtauction.go.kr)의 물건검색 API를 호출해
2026년 주택(주거용 건물)·상업용 건물 경매 물건을 **공고일 기준** 월별로
수집하고,

  1) 엑셀 파일(output/법원경매_주택_2026.xlsx, 법원경매_상업용_2026.xlsx)
  2) 대시보드용 JSON(data.json, data_commercial.json)

으로 저장합니다.

사용법
------
  # 주택 + 상업용 모두, 2026년 1월 ~ 현재 월까지 수집
  python3 fetch_court_auction.py

  # 특정 구분만 수집
  python3 fetch_court_auction.py --category 주택
  python3 fetch_court_auction.py --category 상업용

  # 최신 월(데이터가 있는 가장 최근 월)만 수집
  python3 fetch_court_auction.py --latest

  # 특정 월만 수집
  python3 fetch_court_auction.py --month 2026-06

  # 네트워크 없이 샘플 데이터로 전체 파이프라인 실행(엑셀 + JSON 생성)
  python3 fetch_court_auction.py --sample

필요 패키지: requests, openpyxl  (pip install requests openpyxl)

주의
----
- 법원경매정보 사이트는 해외 IP 접속을 차단하므로 국내 네트워크에서 실행해야 합니다.
- 공공 사이트이므로 요청 간 지연(기본 1초)을 두고 과도한 호출을 하지 않습니다.
- 사이트 개편 시 API 필드명이 바뀔 수 있습니다. 이 경우 브라우저 개발자도구(F12)
  → Network 탭에서 "물건상세검색" 요청을 확인해 아래 API_* 상수를 갱신하세요.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import random
import sys
import time
from pathlib import Path

try:
    import requests
except ImportError:
    requests = None

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"
YEAR = 2026

# 물건 구분별 설정: 용도 중분류 코드와 저장 경로
# (코드값은 사이트 개편 시 개발자도구에서 재확인)
CATEGORIES = {
    "주택": {
        "mcl_code": "20100",   # 중분류: 주거용건물
        "excel": OUTPUT_DIR / "법원경매_주택_2026.xlsx",
        "json": BASE_DIR / "data.json",
    },
    "상업용": {
        "mcl_code": "20200",   # 중분류: 상업용건물
        "excel": OUTPUT_DIR / "법원경매_상업용_2026.xlsx",
        "json": BASE_DIR / "data_commercial.json",
    },
}

# ---------------------------------------------------------------------------
# 법원경매정보 API 설정
# (2024년 개편된 신규 사이트 기준. 개편 시 개발자도구로 재확인 필요)
# ---------------------------------------------------------------------------
API_URL = "https://www.courtauction.go.kr/pgj/pgj15B/selectAuctnGdsDtlSrchRslt.on"
API_REFERER = "https://www.courtauction.go.kr/pgj/PGJ15BM01.jsp"
API_HEADERS = {
    "Content-Type": "application/json; charset=UTF-8",
    "Accept": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    ),
    "Referer": API_REFERER,
}
PAGE_SIZE = 40          # 사이트 기본 페이지 크기
REQUEST_DELAY_SEC = 1.0  # 요청 간 지연(서버 부하 방지)
MAX_RETRY = 3

# 용도 대분류 코드. 사이트 개편 시 개발자도구에서 코드값 재확인.
USAGE_LARGE_CODE = "20000"   # 대분류: 건물

# 응답 JSON에서 항목 리스트가 담기는 후보 키(사이트 버전에 따라 다를 수 있음)
RESULT_LIST_KEYS = ["dlt_srchResult", "dlt_srchRslt", "data"]

# 필드 정규화: 우리 컬럼명 → API 응답 후보 키 목록(앞에서부터 먼저 발견되는 값 사용)
FIELD_CANDIDATES = {
    "사건번호": ["userCsNo", "csNo", "printCsNo"],
    "법원": ["cortOfcNm", "cortNm", "jiwonNm"],
    "물건번호": ["dspslGdsSeq", "maemulSer"],
    "용도": ["dspslGdsLstUsgNm", "sclDspslGdsLstUsgNm", "mclDspslGdsLstUsgNm", "usgNm"],
    "소재지": ["printSt", "rprsLtnoAdrs", "rdnmAdrs", "adrs"],
    "감정가": ["aeeEvlAmt", "aeeWvalAmt", "gamevalAmt"],
    "최저매각가": ["fstPbancLwsDspslPrc", "lwsDspslPrc", "minmaePrice"],
    "유찰횟수": ["flbdNcnt", "yuchalCnt"],
    "공고일": ["dspslPrceNtcYmd", "pbancYmd", "ntcYmd"],
    "매각기일": ["dspslDxdyYmd", "maegakYmd", "dxdyYmd"],
    "상태": ["auctnGdsStatNm", "gdsStatNm", "statNm"],
}


def build_payload(page_no: int, start_ymd: str, end_ymd: str, mcl_code: str) -> dict:
    """월별 공고일 범위 검색 페이로드. 필드명은 사이트 개편 시 재확인 필요."""
    return {
        "dma_pageInfo": {
            "pageNo": page_no,
            "pageSize": PAGE_SIZE,
            "bfPageNo": max(page_no - 1, 1),
            "startRowNo": (page_no - 1) * PAGE_SIZE + 1,
            "totalYn": "Y",
        },
        "dma_srchGdsDtlSrchInfo": {
            "bidDvsCd": "000331",          # 기일입찰
            "mvprpRletDvsCd": "00031R",    # 부동산
            "cortAuctnSrchCondCd": "0004601",
            "lclDspslGdsLstUsgCd": USAGE_LARGE_CODE,
            "mclDspslGdsLstUsgCd": mcl_code,
            "sclDspslGdsLstUsgCd": "",
            "cortOfcCd": "",               # 전체 법원
            "rprsAdongSdCd": "",           # 전체 시도
            "rprsAdongSggCd": "",
            "rprsAdongEmdCd": "",
            # 공고일(게시일) 범위 — 공고일 기준 수집의 핵심 파라미터
            "dspslPrceNtcYmdStart": start_ymd,
            "dspslPrceNtcYmdEnd": end_ymd,
            # 일부 버전은 매각기일 범위 키를 요구하므로 함께 전달(무시되면 무해)
            "dspslDxdyYmdStart": "",
            "dspslDxdyYmdEnd": "",
            "aeeWvalAmtMin": "",
            "aeeWvalAmtMax": "",
            "flbdNcntMin": "",
            "flbdNcntMax": "",
            "pgmId": "PGJ15BM01",
            "menuNm": "물건상세검색",
        },
    }


def pick_first(item: dict, keys: list[str]):
    for k in keys:
        v = item.get(k)
        if v not in (None, ""):
            return v
    return ""


def normalize_record(item: dict) -> dict:
    rec = {col: pick_first(item, keys) for col, keys in FIELD_CANDIDATES.items()}
    for col in ("감정가", "최저매각가", "유찰횟수"):
        try:
            rec[col] = int(str(rec[col]).replace(",", "") or 0)
        except ValueError:
            rec[col] = 0
    for col in ("공고일", "매각기일"):
        s = str(rec[col])
        if len(s) == 8 and s.isdigit():
            rec[col] = f"{s[:4]}-{s[4:6]}-{s[6:]}"
    addr = str(rec["소재지"])
    rec["시도"] = addr.split()[0] if addr else ""
    return rec


def fetch_month(session, year: int, month: int, mcl_code: str) -> list[dict]:
    """한 달치(공고일 기준) 데이터를 페이지네이션하며 전부 수집."""
    start = f"{year}{month:02d}01"
    last_day = (dt.date(year + month // 12, month % 12 + 1, 1) - dt.timedelta(days=1)).day
    end = f"{year}{month:02d}{last_day:02d}"

    records: list[dict] = []
    page = 1
    while True:
        payload = build_payload(page, start, end, mcl_code)
        data = None
        for attempt in range(1, MAX_RETRY + 1):
            try:
                resp = session.post(API_URL, json=payload, headers=API_HEADERS, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                break
            except Exception as e:  # noqa: BLE001
                if attempt == MAX_RETRY:
                    raise RuntimeError(
                        f"{year}-{month:02d} p{page} 요청 실패: {e}\n"
                        "  - 국내 네트워크인지 확인하세요(해외 IP 차단).\n"
                        "  - 사이트 개편으로 API가 바뀌었다면 스크립트 상단 주석을 참고해 갱신하세요.\n"
                        "  - 네트워크 없이 테스트하려면 --sample 옵션을 사용하세요."
                    ) from e
                time.sleep(2 ** attempt)

        items = None
        for key in RESULT_LIST_KEYS:
            if isinstance(data.get(key), list):
                items = data[key]
                break
        if items is None:
            raise RuntimeError(
                f"응답에서 결과 리스트를 찾지 못했습니다. 최상위 키: {list(data)[:10]}\n"
                "RESULT_LIST_KEYS에 실제 키를 추가하세요."
            )

        records.extend(normalize_record(it) for it in items)
        if len(items) < PAGE_SIZE:
            break
        page += 1
        time.sleep(REQUEST_DELAY_SEC)

    print(f"  {year}-{month:02d}: {len(records)}건 수집")
    return records


# ---------------------------------------------------------------------------
# 샘플 데이터 생성 (--sample): 네트워크 없이 파이프라인 전체를 검증
# ---------------------------------------------------------------------------
SIDO_WEIGHTS = [
    ("경기", 26), ("서울", 14), ("부산", 8), ("인천", 7), ("경남", 6),
    ("경북", 5), ("충남", 5), ("전남", 4), ("대구", 4), ("전북", 4),
    ("충북", 4), ("강원", 3), ("광주", 3), ("대전", 3), ("울산", 2),
    ("제주", 1), ("세종", 1),
]
USAGE_WEIGHTS = {
    "주택": [
        ("아파트", 38), ("다세대주택", 22), ("단독주택", 14), ("오피스텔", 9),
        ("연립주택", 8), ("다가구주택", 5), ("주상복합", 4),
    ],
    "상업용": [
        ("근린생활시설", 34), ("상가", 21), ("점포", 13), ("사무실", 12),
        ("숙박시설", 9), ("근린상가", 8), ("목욕탕", 3),
    ],
}
COURTS = {
    "서울": ["서울중앙지방법원", "서울동부지방법원", "서울남부지방법원", "서울북부지방법원", "서울서부지방법원"],
    "경기": ["수원지방법원", "의정부지방법원", "수원지방법원 성남지원", "의정부지방법원 고양지원"],
    "부산": ["부산지방법원", "부산지방법원 동부지원"], "인천": ["인천지방법원", "인천지방법원 부천지원"],
    "대구": ["대구지방법원"], "광주": ["광주지방법원"], "대전": ["대전지방법원"], "울산": ["울산지방법원"],
    "경남": ["창원지방법원", "창원지방법원 진주지원"], "경북": ["대구지방법원 포항지원", "대구지방법원 안동지원"],
    "충남": ["대전지방법원 천안지원", "대전지방법원 서산지원"], "충북": ["청주지방법원"],
    "전남": ["광주지방법원 순천지원", "광주지방법원 목포지원"], "전북": ["전주지방법원", "전주지방법원 군산지원"],
    "강원": ["춘천지방법원", "춘천지방법원 원주지원"], "제주": ["제주지방법원"], "세종": ["대전지방법원"],
}
BASE_PRICE = {  # 용도별 감정가 중심값(만원)
    # 주택
    "아파트": 42000, "주상복합": 55000, "오피스텔": 21000, "단독주택": 38000,
    "다세대주택": 17000, "연립주택": 19000, "다가구주택": 45000,
    # 상업용
    "근린생활시설": 48000, "상가": 35000, "점포": 22000, "사무실": 40000,
    "숙박시설": 90000, "근린상가": 30000, "목욕탕": 65000,
}
SIDO_MULT = {"서울": 2.3, "경기": 1.4, "인천": 1.0, "부산": 1.0, "대구": 0.9, "대전": 0.9,
             "광주": 0.8, "울산": 0.9, "세종": 1.1, "제주": 1.0}


def weighted_choice(rng: random.Random, pairs):
    total = sum(w for _, w in pairs)
    r = rng.uniform(0, total)
    acc = 0
    for v, w in pairs:
        acc += w
        if r <= acc:
            return v
    return pairs[-1][0]


def generate_sample(year: int, upto_month: int, category: str) -> list[dict]:
    # 구분별로 시드를 달리해 서로 다른 분포 생성
    rng = random.Random(20260712 + hash(category) % 1000)
    today = dt.date.today()
    is_commercial = category == "상업용"
    records = []
    seq = 0
    for month in range(1, upto_month + 1):
        # 상업용 물건 수는 주택보다 적게
        n = rng.randint(150, 240) if is_commercial else rng.randint(330, 470)
        last_day = (dt.date(year + month // 12, month % 12 + 1, 1) - dt.timedelta(days=1)).day
        # 현재 진행 중인 월은 오늘까지의 공고만 생성
        if year == today.year and month == today.month:
            last_day = today.day
            n = int(n * today.day / 30)
        for _ in range(n):
            seq += 1
            sido = weighted_choice(rng, SIDO_WEIGHTS)
            usage = weighted_choice(rng, USAGE_WEIGHTS[category])
            court = rng.choice(COURTS[sido])
            appraisal = int(BASE_PRICE[usage] * SIDO_MULT.get(sido, 0.75)
                            * rng.lognormvariate(0, 0.55 if is_commercial else 0.45)) * 10000
            appraisal = max(appraisal // 1000000 * 1000000, 30000000)
            # 상업용은 유찰이 더 잦은 경향
            fails = min(int(rng.expovariate(0.8 if is_commercial else 1.1)), 5)
            low_price = int(appraisal * (0.8 ** fails) // 100000 * 100000)
            notice_day = rng.randint(1, last_day)
            notice = dt.date(year, month, notice_day)
            sale = notice + dt.timedelta(days=rng.randint(14, 35))
            status = "신건" if fails == 0 else "유찰"
            case_base = 200000 if is_commercial else 100000
            records.append({
                "사건번호": f"{year}타경{case_base + seq}",
                "법원": court,
                "물건번호": 1,
                "용도": usage,
                "소재지": f"{sido} (샘플) 주소 {seq}",
                "감정가": appraisal,
                "최저매각가": low_price,
                "유찰횟수": fails,
                "공고일": notice.isoformat(),
                "매각기일": sale.isoformat(),
                "상태": status,
                "시도": sido,
            })
    return records


# ---------------------------------------------------------------------------
# 저장: 엑셀 + JSON
# ---------------------------------------------------------------------------
COLUMNS = ["사건번호", "법원", "물건번호", "용도", "시도", "소재지",
           "감정가", "최저매각가", "유찰횟수", "공고일", "매각기일", "상태"]


def save_excel(records: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    def write_sheet(ws, rows):
        ws.append(COLUMNS)
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(c, "") for c in COLUMNS])
        for idx, col in enumerate(COLUMNS, start=1):
            letter = get_column_letter(idx)
            width = {"소재지": 34, "법원": 22, "사건번호": 16, "감정가": 14,
                     "최저매각가": 14, "공고일": 12, "매각기일": 12}.get(col, 10)
            ws.column_dimensions[letter].width = width
        money_cols = [COLUMNS.index("감정가") + 1, COLUMNS.index("최저매각가") + 1]
        for row in ws.iter_rows(min_row=2):
            for ci in money_cols:
                row[ci - 1].number_format = "#,##0"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws_all = wb.active
    ws_all.title = "전체"
    write_sheet(ws_all, records)

    months = sorted({r["공고일"][:7] for r in records if r.get("공고일")})
    for m in months:
        ws = wb.create_sheet(m.replace("-", "년 ") + "월")
        write_sheet(ws, [r for r in records if str(r.get("공고일", "")).startswith(m)])

    ws_sum = wb.create_sheet("월별요약", 1)
    ws_sum.append(["공고월", "물건수", "평균 감정가", "평균 최저매각가", "평균 유찰횟수"])
    for c in range(1, 6):
        cell = ws_sum.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for m in months:
        rows = [r for r in records if str(r.get("공고일", "")).startswith(m)]
        n = len(rows)
        ws_sum.append([
            m, n,
            round(sum(r["감정가"] for r in rows) / n) if n else 0,
            round(sum(r["최저매각가"] for r in rows) / n) if n else 0,
            round(sum(r["유찰횟수"] for r in rows) / n, 2) if n else 0,
        ])
    for letter, width in zip("ABCDE", (10, 10, 16, 16, 14)):
        ws_sum.column_dimensions[letter].width = width
    for row in ws_sum.iter_rows(min_row=2):
        row[2].number_format = "#,##0"
        row[3].number_format = "#,##0"

    wb.save(path)
    print(f"엑셀 저장: {path} ({len(records)}건, 시트 {len(wb.sheetnames)}개)")


def save_json(records: list[dict], path: Path, source: str, category: str) -> None:
    months = sorted({r["공고일"][:7] for r in records if r.get("공고일")})
    payload = {
        "meta": {
            "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
            "source": source,
            "year": YEAR,
            "category": category,
            "months": months,
            "count": len(records),
            "basis": "공고일",
        },
        "records": [{k: r.get(k, "") for k in COLUMNS} for r in records],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"JSON 저장: {path} ({len(records)}건)")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def run_category(category: str, args, current_month: int, session) -> bool:
    cfg = CATEGORIES[category]

    if args.sample:
        print(f"[{category}] 샘플 모드: {YEAR}년 1월~{current_month}월 데이터 생성")
        records = generate_sample(YEAR, current_month, category)
        source = "샘플 데이터 (--sample)"
    else:
        if args.month:
            y, m = args.month.split("-")
            months = [(int(y), int(m))]
        elif args.latest:
            months = [(YEAR, mm) for mm in range(current_month, 0, -1)]
        else:
            months = [(YEAR, mm) for mm in range(1, current_month + 1)]

        records = []
        print(f"[{category}] 법원경매정보 수집 시작 (공고일 기준)")
        for y, m in months:
            month_records = fetch_month(session, y, m, cfg["mcl_code"])
            records.extend(month_records)
            if args.latest and month_records:
                print(f"[{category}] 최신 데이터 월: {y}-{m:02d}")
                break
            time.sleep(REQUEST_DELAY_SEC)
        source = "법원경매정보(courtauction.go.kr)"

    if not records:
        print(f"[{category}] 수집된 데이터가 없습니다.", file=sys.stderr)
        return False

    records.sort(key=lambda r: (r.get("공고일", ""), r.get("사건번호", "")))
    save_excel(records, cfg["excel"])
    save_json(records, cfg["json"], source, category)

    months = sorted({r["공고일"][:7] for r in records})
    print(f"\n[{category}] 월별 요약")
    for m in months:
        rows = [r for r in records if r["공고일"].startswith(m)]
        avg = sum(r["감정가"] for r in rows) / len(rows)
        print(f"  {m}: {len(rows):>5}건, 평균 감정가 {avg/1e8:.2f}억원")
    print()
    return True


def main() -> int:
    ap = argparse.ArgumentParser(
        description="법원경매 2026년 월별 부동산(주택/상업용) 데이터 수집기(공고일 기준)")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--month", help="특정 월만 수집 (예: 2026-06)")
    g.add_argument("--latest", action="store_true", help="데이터가 있는 최신 월만 수집")
    ap.add_argument("--sample", action="store_true", help="네트워크 없이 샘플 데이터 생성")
    ap.add_argument("--category", choices=[*CATEGORIES, "전체"], default="전체",
                    help="물건 구분 (기본: 전체 = 주택 + 상업용)")
    args = ap.parse_args()

    today = dt.date.today()
    current_month = today.month if today.year == YEAR else (12 if today.year > YEAR else 1)

    session = None
    if not args.sample:
        if requests is None:
            print("requests 패키지가 없습니다: pip install requests openpyxl", file=sys.stderr)
            return 1
        session = requests.Session()
        # 첫 페이지 접속으로 세션 쿠키 확보
        try:
            session.get("https://www.courtauction.go.kr/pgj/index.on",
                        headers=API_HEADERS, timeout=30)
        except Exception as e:  # noqa: BLE001
            print(f"사이트 접속 실패: {e}\n국내 네트워크에서 실행하거나 --sample을 사용하세요.",
                  file=sys.stderr)
            return 1

    categories = list(CATEGORIES) if args.category == "전체" else [args.category]
    ok = all([run_category(c, args, current_month, session) for c in categories])
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
